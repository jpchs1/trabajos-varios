#!/usr/bin/env python3
"""Planilla del itinerario propuesto para mandarle al operador.

    python3 build-excel.py

Tres hojas: el itinerario con dos columnas en blanco para que el operador
responda si es factible, las preguntas transversales, y una hoja de referencia
con lo que cambió y los valores cotizados.

El itinerario se lee de itinerario.mjs y los precios de contenido.mjs, las
mismas fuentes que los PDF: acá no hay horarios ni montos escritos de nuevo.
"""

import json
import subprocess
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AQUI = Path(__file__).parent

def desde_node(expr):
    out = subprocess.run(['node', '--input-type=module', '-e',
                          f"{expr}.then(m=>console.log(JSON.stringify(m)))"],
                         cwd=AQUI, capture_output=True, text=True, check=True)
    return json.loads(out.stdout)

COT = desde_node("import('./contenido.mjs').then(m=>({doc:m.doc,cliente:m.cliente,programa:m.programa}))")
ITI = desde_node("import('./itinerario.mjs').then(m=>({dias:m.DIAS,cambios:m.CAMBIOS,generales:m.PREGUNTAS_GENERALES}))")
DOC, CLIENTE, PROGRAMA = COT['doc'], COT['cliente'], COT['programa']
DIAS, CAMBIOS, GENERALES = ITI['dias'], ITI['cambios'], ITI['generales']
PAX = DOC['pax']

# Ningún servicio cotizado puede quedar fuera del itinerario. Puritama se
# escapó una vez de la columna de confirmaciones; esto lo hace imposible.
CUBIERTOS = [i for d in DIAS for b in d['blks'] for i in b.get('cot', [])]
FALTAN = [s['id'] for s in PROGRAMA if s['id'] not in CUBIERTOS]
SOBRAN = [i for i in CUBIERTOS if i not in {s['id'] for s in PROGRAMA}]
REPETIDOS = [i for i in set(CUBIERTOS) if CUBIERTOS.count(i) > 1]
SIN_CONF = [b['t']['es'] for d in DIAS for b in d['blks'] if b['tipo'] == 'tourevo' and not b.get('conf')]
if FALTAN or SOBRAN or REPETIDOS or SIN_CONF:
    raise SystemExit(
        f'El itinerario no cuadra con la cotización.\n'
        f'  Servicios cotizados que no aparecen: {FALTAN or "ninguno"}\n'
        f'  Ids del itinerario que no existen en la cotización: {SOBRAN or "ninguno"}\n'
        f'  Ids repetidos: {REPETIDOS or "ninguno"}\n'
        f'  Servicios sin texto de confirmación: {SIN_CONF or "ninguno"}')

# Dónde cae cada servicio cotizado dentro del itinerario.
CUANDO = {}

# --- Estilos ------------------------------------------------------------------

F = 'Arial'
CABECERA, TINTA, SUAVE = 'FF0B5B54', 'FF16232B', 'FF566A72'
ACENTO, AVISO = 'FF0F766E', 'FFA0561F'
LINEA, GRIS, VERDE, AMBAR, AMARILLO = 'FFD3E0E0', 'FFF3F8F8', 'FFEAF5F3', 'FFFDF3E9', 'FFFFF9DB'

th    = Font(name=F, size=9,  bold=True, color='FFFFFFFF')
td    = Font(name=F, size=10, color=TINTA)
td_b  = Font(name=F, size=10, bold=True, color=TINTA)
td_sv = Font(name=F, size=10, color=SUAVE)
td_ac = Font(name=F, size=10, bold=True, color=ACENTO)
td_av = Font(name=F, size=10, bold=True, color=AVISO)
td_it = Font(name=F, size=9,  italic=True, color=SUAVE)

fill_th, fill_gris = PatternFill('solid', fgColor=CABECERA), PatternFill('solid', fgColor=GRIS)
fill_verd, fill_amb = PatternFill('solid', fgColor=VERDE), PatternFill('solid', fgColor=AMBAR)
fill_am = PatternFill('solid', fgColor=AMARILLO)

borde = Border(*[Side(style='thin', color=LINEA)] * 4)
wrap  = Alignment(vertical='top', wrap_text=True)
centro= Alignment(vertical='top', horizontal='center')
sinsalto = Alignment(vertical='center', wrap_text=True)

def hoja(ws, cols, fila_th):
    for i, (titulo, ancho) in enumerate(cols, start=1):
        c = ws.cell(row=fila_th, column=i, value=titulo)
        c.font, c.fill, c.border, c.alignment = th, fill_th, borde, sinsalto
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[fila_th].height = 26
    ws.freeze_panes = ws.cell(row=fila_th + 1, column=1)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f'{fila_th}:{fila_th}'
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

def poner(ws, fila, valores, fill=None, fuente=None):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=fila, column=i, value=v)
        c.font, c.alignment, c.border = fuente or td, wrap, borde
        if fill:
            c.fill = fill
    return fila + 1

def sin_html(t):
    return t.replace('<b>', '').replace('</b>', '') if t else ''

DIA_ES = ['lunes','martes','miércoles','jueves','viernes','sábado','domingo']
def dia_nombre(iso):
    y, m, d = (int(x) for x in iso.split('-'))
    return f'{DIA_ES[date(y, m, d).weekday()]} {d}'

OPERA = {'tourevo': 'Tourevo', 'cliente': 'Cliente', 'libre': 'Libre'}

wb = Workbook()

# --- Hoja 1: itinerario, con dos columnas para que el operador responda -------

ws = wb.active
ws.title = 'Itinerario'
ws['A1'] = f'{CLIENTE["nombre"]} · San Pedro de Atacama · 22 al 26 de diciembre de 2026 · {PAX} pasajeros'
ws['A1'].font = Font(name=F, size=13, bold=True, color=TINTA)
ws['A2'] = ('PARA EL OPERADOR: complete sólo las columnas I («¿Factible?») y J («Comentarios»), en amarillo. '
            'El resto es referencia. En «¿Factible?» ponga Sí, No o Con cambios. '
            'La columna Nº numera los 8 servicios que operamos nosotros: las filas sin número son vuelos, comidas o tiempo libre.')
ws['A2'].font = Font(name=F, size=10, color=AVISO)
ws.merge_cells('A1:J1')
ws.merge_cells('A2:J2')
ws.row_dimensions[2].height = 16

COLS = [('Nº',6),('Fecha',11),('Día',13),('Inicio',8),('Fin',8),('Servicio',52),('Opera',10),
        ('Qué necesitamos confirmar',60),('¿Factible?',13),('Comentarios del operador',40)]
hoja(ws, COLS, 4)

fila = poner(ws, 5, ['0','2026-12-00','ejemplo','00:00','00:00','EJEMPLO — borre esta fila','Tourevo',
                     'Así se ve una línea llena.','Con cambios','Se puede, pero salida 10:00.'],
             fill=fill_am, fuente=td_it)

FILL = {'movido': fill_amb, 'nuevo': fill_verd}
fill_num = PatternFill('solid', fgColor=CABECERA)
primera = fila
nro = 0
for d in DIAS:
    for b in d['blks']:
        es_nuestro = b['tipo'] == 'tourevo'
        if es_nuestro:
            nro += 1
            for i in b.get('cot', []):
                CUANDO[i] = f"{dia_nombre(d['f'])} · {b['h']} – {b.get('f2') or ''}"
        f = FILL.get(b.get('e')) if es_nuestro else fill_gris
        servicio = b['t']['es']
        if b.get('n'):
            servicio += '\n' + sin_html(b['n']['es'])
        fila = poner(ws, fila, [
            nro if es_nuestro else None, d['f'], dia_nombre(d['f']), b['h'], b.get('f2') or '',
            servicio, OPERA[b['tipo']], sin_html(b.get('conf', {}).get('es', '')), None, None,
        ], fill=f, fuente=td if es_nuestro else td_sv)
        r = fila - 1
        ws.cell(row=r, column=6).font = td_b if es_nuestro else td_sv
        for col in (4, 5, 7):
            ws.cell(row=r, column=col).alignment = centro
        for col in (9, 10):
            ws.cell(row=r, column=col).fill = fill_am
        # El número va en teal sólido: los ocho servicios nuestros se cuentan de
        # un vistazo y no hay forma de saltarse uno.
        c = ws.cell(row=r, column=1)
        c.alignment = Alignment(vertical='center', horizontal='center')
        if es_nuestro:
            c.fill, c.font = fill_num, Font(name=F, size=11, bold=True, color='FFFFFFFF')
ULTIMA = fila - 1
ws.auto_filter.ref = f'A4:J{ULTIMA}'
TOTAL_SERVICIOS = nro

# --- Hoja 2: preguntas transversales ------------------------------------------

ws2 = wb.create_sheet('Preguntas')
ws2['A1'] = 'Preguntas que no cuelgan de una línea del itinerario'
ws2['A1'].font = Font(name=F, size=13, bold=True, color=TINTA)
ws2['A2'] = 'Complete sólo la columna D («Respuesta del operador»), en amarillo.'
ws2['A2'].font = Font(name=F, size=10, color=AVISO)
hoja(ws2, [('Nº',6),('Pregunta',46),('Detalle',72),('Respuesta del operador',40)], 4)
fila = poner(ws2, 5, [0, 'EJEMPLO — borre esta fila', 'Así se ve una línea llena.', 'Sí, sin recargo.'],
             fill=fill_am, fuente=td_it)
for i, q in enumerate(GENERALES, start=1):
    fila = poner(ws2, fila, [i, q['q']['es'], q['d']['es'], None])
    ws2.cell(row=fila-1, column=1).alignment = centro
    ws2.cell(row=fila-1, column=2).font = td_b
    ws2.cell(row=fila-1, column=4).fill = fill_am

# --- Hoja 3: referencia -------------------------------------------------------

ws3 = wb.create_sheet('Referencia')
ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 42
ws3.column_dimensions['C'].width = 88
ws3['A1'] = f'{DOC["numero"]} · referencia interna'
ws3['A1'].font = Font(name=F, size=13, bold=True, color=TINTA)

r = 3
ws3.cell(row=r, column=1, value='QUÉ CAMBIÓ RESPECTO DE LO QUE EL CLIENTE YA VIO').font = Font(name=F, size=10, bold=True, color=CABECERA)
r += 1
for i, c in enumerate(CAMBIOS, start=1):
    ws3.cell(row=r, column=1, value=i).font = td_b
    ws3.cell(row=r, column=1).alignment = centro
    ws3.cell(row=r, column=2, value=c['tit']['es']).font = td_b
    ws3.cell(row=r, column=2).alignment = wrap
    cc = ws3.cell(row=r, column=3, value=' '.join(sin_html(p) for p in c['cuerpo']['es']))
    cc.font, cc.alignment = td, wrap
    r += 1

r += 1
ws3.cell(row=r, column=1, value='VALORES COTIZADOS').font = Font(name=F, size=10, bold=True, color=CABECERA)
r += 1
TH3 = r
ANCHOS3 = [6, 42, 30, 15, 15, 15, 15]
for i, t in enumerate(['Nº', 'Servicio cotizado', 'Cuándo va en el itinerario', 'Servicio p/p',
                       'Entradas p/p', 'Subtotal p/p', f'Total {PAX} pax'], start=1):
    c = ws3.cell(row=TH3, column=i, value=t)
    c.font, c.fill, c.border, c.alignment = th, fill_th, borde, sinsalto
    ws3.column_dimensions[get_column_letter(i)].width = ANCHOS3[i-1]

r = TH3 + 1
for i, sv in enumerate(PROGRAMA, start=1):
    poner(ws3, r, [i, sv['titulo']['es'], CUANDO[sv['id']], sv['valor'], sv['entradas'], None, None])
    ws3.cell(row=r, column=1).alignment = centro
    ws3.cell(row=r, column=3).font = td_ac
    ws3.cell(row=r, column=6, value=f'=D{r}+E{r}')
    ws3.cell(row=r, column=7, value=f'=F{r}*{PAX}')
    for col in range(4, 8):
        cc = ws3.cell(row=r, column=col)
        cc.number_format = '#,##0'
        cc.alignment = Alignment(vertical='top', horizontal='right')
        cc.font = td_ac if col == 6 else td
        cc.border = borde
    r += 1

poner(ws3, r, ['', 'TOTAL DEL PACK', '', None, None, None, None], fill=fill_verd)
for col, letra in ((4,'D'), (5,'E'), (6,'F'), (7,'G')):
    c = ws3.cell(row=r, column=col, value=f'=SUM({letra}{TH3+1}:{letra}{r-1})')
    c.number_format, c.font = '#,##0', Font(name=F, size=10, bold=True, color=CABECERA)
    c.alignment, c.fill, c.border = Alignment(vertical='center', horizontal='right'), fill_verd, borde
ws3.cell(row=r, column=2).font = Font(name=F, size=10, bold=True, color=CABECERA)
TOT = r

r += 2
ws3.cell(row=r, column=1, value='RESUMEN Y FUENTES').font = Font(name=F, size=10, bold=True, color=CABECERA)
r += 1
RESUMEN = [
 ('Bloques del itinerario', f'=COUNTA(Itinerario!B6:B{ULTIMA})'),
 ('De ellos, servicios que operamos', f'=COUNTA(Itinerario!A6:A{ULTIMA})'),
 ('Ítems cotizados que cubren', f'=COUNTA(B{TH3+1}:B{TOT-1})'),
 ('Total del pack por persona', f'=F{TOT}'),
 (f'Total del pack para {PAX} pax', f'=G{TOT}'),
]
for k, f in RESUMEN:
    ws3.cell(row=r, column=2, value=k).font = td
    c = ws3.cell(row=r, column=3, value=f)
    c.font, c.number_format = td_b, '#,##0'
    c.alignment = Alignment(vertical='top', horizontal='left')
    r += 1

r += 1
FUENTES = [
 ('Itinerario', 'itinerario.mjs, la misma fuente que el PDF del itinerario propuesto. Horarios y notas no se escriben dos veces.'),
 ('Los 9 ítems en 8 bloques', 'Quitor y el Valle de la Muerte van juntos en un solo medio día el 24, así que los nueve servicios cotizados caben en ocho bloques. El build falla si alguno queda fuera.'),
 ('Valores', f'contenido.mjs, la misma fuente que la cotización {DOC["numero"]}. Sujetos a recotización: astronomía privada, alcance del full day del 25 y el medio día combinado Quitor + Valle de la Muerte.'),
 ('Cejar cierra los martes', 'Dato del cliente. El único martes del viaje es el 22, día de llegada sin excursiones, así que no afecta.'),
 ('Fase lunar del 24', 'Cálculo propio sobre la luna nueva del 6 ene 2000: la noche del 24 va al 99,4% de iluminación y ninguna otra noche del viaje baja del 92%.'),
 ('Tiempos de camino', 'Estimados: San Pedro a Cejar unos 35 minutos, Quitor a Valle de la Muerte unos 20. Los confirma el operador.'),
]
for k, v in FUENTES:
    ws3.cell(row=r, column=2, value=k).font = td_b
    c = ws3.cell(row=r, column=3, value=v); c.font, c.alignment = td, wrap
    r += 1

SALIDA = AQUI / f'Tourevo-{DOC["numero"]}-{CLIENTE["nombre"]}-Itinerario-para-operador.xlsx'
wb.save(SALIDA)
print(f'✓ {SALIDA.name} · {ULTIMA-primera+1} bloques · {TOTAL_SERVICIOS} servicios nuestros · '
      f'{len(PROGRAMA)} ítems cotizados, todos ubicados · {len(GENERALES)} preguntas')
